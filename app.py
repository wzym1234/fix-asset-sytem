# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, session, send_file, redirect, url_for
import pandas as pd
import mysql.connector
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional
import datetime
import os

# Initialize Flask application
app = Flask(__name__, static_folder='static')
# Set a secret key for session management
app.secret_key = 'your_secret_key_here'  

# Database configuration
DB_CONFIG = {
    'user': 'your_user',
    'password': 'your_password',
    'host': '10.132.1.10',
    'database': 'pandian'
}

# Configure local path
BASE_PATH = 'data/'

# Create a database connection
def create_db_connection():
    return mysql.connector.connect(**DB_CONFIG)

# Create tables
def create_tables():
    conn = create_db_connection()
    cursor = conn.cursor()
    # Create user table
    create_user_table = """
    CREATE TABLE IF NOT EXISTS users (
        id INT AUTO_INCREMENT PRIMARY KEY,
        username VARCHAR(255),
        password VARCHAR(255)
    )
    """
    cursor.execute(create_user_table)
    # Create assets table with user ID field
    create_asset_table = """
    CREATE TABLE IF NOT EXISTS assets (
        id INT AUTO_INCREMENT PRIMARY KEY,
        user_id INT,
        设备编码 VARCHAR(255),
        设备名称 VARCHAR(255),
        型号 VARCHAR(255),
        品牌名称 VARCHAR(255),
        设备类别编码 VARCHAR(255),
        设备类别名称 VARCHAR(255),
        状态名称 VARCHAR(255),
        管理人 VARCHAR(255),
        管理部门 VARCHAR(255),
        责任人 VARCHAR(255),
        使用部门 VARCHAR(255),
        位置 VARCHAR(255),
        盘点状态 VARCHAR(255),
        FOREIGN KEY (user_id) REFERENCES users(id)
    )
    """
    cursor.execute(create_asset_table)
    # Create inventory issues table with user ID field
    create_issue_table = """
    CREATE TABLE IF NOT EXISTS inventory_issues (
        id INT AUTO_INCREMENT PRIMARY KEY,
        user_id INT,
        设备编码 VARCHAR(255),
        记录时间 DATETIME,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )
    """
    cursor.execute(create_issue_table)
    conn.commit()
    cursor.close()
    conn.close()

# Clear the assets and inventory issues tables for the current user
def clear_tables(user_id):
    conn = create_db_connection()
    cursor = conn.cursor()
    # Clear assets table
    clear_assets_query = "DELETE FROM assets WHERE user_id = %s"
    cursor.execute(clear_assets_query, (user_id,))
    # Clear inventory issues table
    clear_issues_query = "DELETE FROM inventory_issues WHERE user_id = %s"
    cursor.execute(clear_issues_query, (user_id,))
    conn.commit()
    cursor.close()
    conn.close()

# Insert data into the database
def insert_data_to_db(df, user_id):
    conn = create_db_connection()
    cursor = conn.cursor()
    for _, row in df.iterrows():
        values = (user_id,) + tuple(row)
        insert_query = f"INSERT INTO assets (user_id, {', '.join(df.columns)}) VALUES ({','.join(['%s'] * len(values))})"
        cursor.execute(insert_query, values)
    conn.commit()
    cursor.close()
    conn.close()

# Retrieve asset data for the current user from the database
def get_asset_data(user_id):
    conn = create_db_connection()
    cursor = conn.cursor(dictionary=True)
    query = "SELECT * FROM assets WHERE user_id = %s"
    cursor.execute(query, (user_id,))
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results

# Update the inventory status for the current user
def update_inventory_status(codes, user_id):
    conn = create_db_connection()
    cursor = conn.cursor()
    placeholders = ', '.join(['%s'] * len(codes))
    update_query = f"UPDATE assets SET 盘点状态 = '已盘点' WHERE 设备编码 IN ({placeholders}) AND user_id = %s"
    codes = tuple(codes) + (user_id,)
    cursor.execute(update_query, codes)
    conn.commit()
    cursor.close()
    conn.close()

# Record equipment codes with physical presence but no accounting records for the current user
def record_inventory_issue(code, user_id):
    conn = create_db_connection()
    cursor = conn.cursor()
    insert_query = "INSERT INTO inventory_issues (user_id, 设备编码, 记录时间) VALUES (%s, %s, %s)"
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute(insert_query, (user_id, code, now))
    conn.commit()
    cursor.close()
    conn.close()

# Retrieve inventory issue data for the current user from the database
def get_inventory_issues(user_id):
    conn = create_db_connection()
    cursor = conn.cursor(dictionary=True)
    query = "SELECT * FROM inventory_issues WHERE user_id = %s"
    cursor.execute(query, (user_id,))
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results

# Export the current user's data to an Excel file from the database
def export_data_to_excel(user_id):
    # Retrieve asset data
    data = get_asset_data(user_id)
    df = pd.DataFrame(data)
    # Retrieve inventory issue data
    issue_data = get_inventory_issues(user_id)
    df_issues = pd.DataFrame(issue_data)

    wb = Workbook()
    # Asset data worksheet
    ws = wb.active
    ws.title = '资产数据'
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(r, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Inventory issues worksheet
    if not df_issues.empty:
        ws_issues = wb.create_sheet('有实物无账务')
        for r_idx, r in enumerate(dataframe_to_rows(df_issues, index=False, header=True), start=1):
            for c_idx, value in enumerate(r, start=1):
                ws_issues.cell(row=r_idx, column=c_idx, value=value)

    filename = f'exported_inventory_{user_id}.xlsx'
    wb.save(filename)
    return filename

KEY_COLUMN = '设备编码'
DISPLAY_FIELDS = [
    '设备编码', '设备名称', '型号', '品牌名称',
    '设备类别编码', '设备类别名称', '状态名称',
    '管理人', '管理部门', '责任人', '使用部门', '位置', '盘点状态'
]

# New route for exporting files
@app.route('/export', methods=['GET'])
def export_file():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']
    filename = export_data_to_excel(user_id)
    try:
        return send_file(filename, as_attachment=True)
    except FileNotFoundError:
        return f"File {filename} not found", 404

# Login page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        conn = create_db_connection()
        cursor = conn.cursor(dictionary=True)
        query = "SELECT id FROM users WHERE username = %s AND password = %s"
        cursor.execute(query, (username, password))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        if user:
            session['user_id'] = user['id']
            return redirect(url_for('manage_inventory'))
        else:
            return render_template('login.html', error='Incorrect username or password')
    return render_template('login.html')

# Logout route
@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('login'))

# Frontend interaction logic
@app.route('/', methods=['GET', 'POST'])
def manage_inventory():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_id = session['user_id']
    results: List[Dict] = []
    error: Optional[str] = None
    success: Optional[str] = None
    search_term = ""
    files = []

    # Retrieve asset data and handle missing columns for empty DataFrame
    data = get_asset_data(user_id)
    df = pd.DataFrame(data)
    
    # Critical fix: Ensure '盘点状态' column exists (handle empty DataFrame scenario)
    if '盘点状态' not in df.columns:
        df['盘点状态'] = ''  # Add empty column when no data

    # Calculate statistics (now df must contain '盘点状态' column)
    total_count = len(df)
    checked_count = len(df[df['盘点状态'] == '已盘点'])
    unchecked_count = len(df[df['盘点状态'] == ''])

    if request.method == 'POST':
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename.endswith('.xlsx'):
                try:
                    # Clear the assets and inventory issues tables for the current user
                    clear_tables(user_id)

                    df = pd.read_excel(file, engine='openpyxl', dtype=str, na_values=['nan', 'NaN', ''])
                    df.columns = df.columns.str.strip()  # Trim whitespace from column names

                    # Check for missing required columns
                    missing_columns = [field for field in DISPLAY_FIELDS if field not in df.columns]
                    for field in missing_columns:
                        if field != '盘点状态':
                            raise ValueError(f"Missing required column in file: {field}")
                        else:
                            df[field] = ''  # Add empty inventory status column

                    # Ensure DataFrame column order matches DISPLAY_FIELDS
                    df = df[DISPLAY_FIELDS]
                    insert_data_to_db(df, user_id)
                    success = f"File {file.filename} uploaded and saved to database successfully"
                except Exception as e:
                    error = str(e)
            else:
                error = "Please upload a valid Excel file (.xlsx format)"

        # Handle search request
        elif 'search_term' in request.form:
            search_term = request.form.get('search_term', '').strip()
            if not search_term:
                error = "Please enter a device code for query"
            else:
                data = get_asset_data(user_id)
                df = pd.DataFrame(data)
                # Check '盘点状态' column again (avoid column missing due to data changes during search)
                if '盘点状态' not in df.columns:
                    df['盘点状态'] = ''
                matched = df[df[KEY_COLUMN].str.upper().str.endswith(search_term)]
                if matched.empty:
                    error = f"No device code ending with [{search_term}] found. Mark as physical presence without accounting records?"
                else:
                    results = matched[DISPLAY_FIELDS].fillna('').to_dict('records')

        # Handle batch marking
        elif 'marked_codes' in request.form:
            marked_codes = request.form.getlist('marked_codes')
            try:
                update_inventory_status(marked_codes, user_id)
                success = f"Successfully marked {len(marked_codes)} records as checked"
            except Exception as e:
                error = str(e)

        # Handle inventory issue records
        elif 'unmatched_code' in request.form:
            code = request.form['unmatched_code']
            action = request.form.get('action', 'no')
            if action == 'yes':
                record_inventory_issue(code, user_id)
                success = f"Device code {code} recorded as physical presence without accounting records"

    return render_template('search.html', results=results, error=error, success=success, search_term=search_term,
                           total_count=total_count, checked_count=checked_count, unchecked_count=unchecked_count)

if __name__ == '__main__':
    create_tables()
    app.run(debug=True, host='0.0.0.0', port=5000)